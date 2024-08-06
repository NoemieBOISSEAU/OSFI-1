# -*- coding: utf-8 -*-
"""
Created on Thu Jul 11 11:18:30 2024

@author: sacha.mailler
"""
import sys
import os
import json
import openpyxl
import statsmodels.api as sm
import time
def return_line(txt,n=52):
    T="\n"
    len_line=0
    for i in range(len(txt)) :
        if i==0 :
            T="\n"
        elif len_line==52 :
            T+="\n"
            len_line=0
        if not(len_line==0 and txt[i]=="\n"):
            if txt[i]=="\n" :
                len_line=0
                T+=txt[i]
            elif txt[i]=="\t":
                T+=" "*4
                len_line+=4
            else :
                T+=txt[i]
                len_line+=1
    return T
#Les objets progression servent à afficher des barres de progressions plutôt  #
#Interactives                                                                 #
###############################################################################
class progression:
    #Initialise l'affichage de la barre de progression :                      #
    #   name : Le nom à afficher au dessus de la barre de progression         #
    ###########################################################################
    def __init__(self,name):
        print(return_line(name,n=52))
        sys.stdout.write(f"\r{' '*52}\r")
        sys.stdout.flush()
        self.progression=0
        self.position=0
        self.max,self.min=100,0
        self.time=0
    #Actuallise le chargement de la barre de progression                      #
    #   progression : pourcentage de progression                              #
    ###########################################################################
    def actualize(self,progression):
        #correction self.min et self.max
        self.min,self.max=max(min(min(self.min,self.max),100),0),max(min(max(self.min,self.max),100),0)
        if progression < 0 :
            progression=0
        if progression > 100 :
            progression=100
        if self.min==self.max :
            progression=self.min
        else :
            progression=self.min+(self.max-self.min)*progression/100
        n=int(progression//2)
        if n==self.progression :
            self.position=(self.position+1)%4
        self.progression=n
        if progression < 100 :
            if self.position==0 :
                car='-'
            if self.position==1:
                car='\\'
            if self.position==2:
                car='|'
            if self.position==3 :
                car='/'
        else :
            car=""
        sys.stdout.flush()
        sys.stdout.write(f"\r{' '*52}\r")
        sys.stdout.write("["+("#"*n)+car+(" "*(49-n))+"]")
        sys.stdout.flush()
    def alert(self,Message):
        print("\n"+Message)
        


class OSFI_regress:
    def __init__(self,path):
        self.path=path
        self.__params={
            "batiment" : None,
        	"code sit RT":None,
        	"code bat RT": None,
        	"RT code sit":None,
        	"RT code bat":None,
        	"RT surf":None,
        	"date":None,
            "consommation chaud" : None,
            "consommation froid" : None,
            "consommation non thermique" : None,
            "Surface" : None,
            "DJU" : None,
            "DJF" : None,
            "typologie" : None,
            "filtres simples" :None,
        	"specific_months":None,
        	"typologies autorisees":None,
        	"verif dqc" :None,
        	"dqc to osfi" : None
        }
        self.__params_num={}
        self.to_remove_dqc=[]
        self.dqc_done=False
        self.excel=None
        self.loaded=False
    def read_json_param(self,name):
        # displayer = progression("Lecture des paramètres dans "+name)
        # displayer.actualize(0)
        with open(os.path.join(self.path,name), 'r', encoding='utf-8') as file :
            temp_dict = json.load(file)
        # displayer.actualize(50)
        for x in self.__params :
            self.__params[x]=temp_dict[x]
        # displayer.actualize(100)
    def add_dqc(self,xls_path):
        self.dqc_done=True
        self.to_remove_dqc=[]
        dqcxlsx = openpyxl.load_workbook(xls_path)
        def get_index(value,sheet):
            flag_found,flag_end=False,False
            i=0
            while not(flag_found or flag_end) :
                i+=1
                if (dqcxlsx[sheet].cell(row=1,column = i).value)==value :
                    flag_found=True
                if (dqcxlsx[sheet].cell(row=1,column = i).value) in [None,"",'""'] :
                    flag_end=True
            if flag_found :
                return i
            return -1
        def get_list(sheet,L_column) :
            count=0
            line=1
            result=[]
            while count<10 :
                line+=1
                flag_void,L=True,[]
                for col in L_column :
                    L.append(dqcxlsx[sheet].cell(row=line,column = col).value)
                    if not(L[-1] in [None,"",'""']) :
                        flag_void = False
                if flag_void:
                    count+=1
                else :
                    count=0
                    result.append(L)
            return result
        for sheet in self.__params["verif dqc"] :
            L_col,L_names=[],[]
            for col in self.__params["verif dqc"][sheet] :
                L_col.append(get_index(col,sheet))
                L_names.append(col)
            rem = get_list(sheet,L_col)
            for x in rem :
                rem2 = {}
                for i in range(len(x)):
                    if L_names[i] in self.__params["dqc to osfi"] :
                        rem2[self.__params["dqc to osfi"][L_names[i]]]=x[i]
                    else :
                        rem2[L_names[i]]=x[i]
                if not(rem2 in self.to_remove_dqc) :
                    self.to_remove_dqc.append(rem2)
        dqcxlsx.close()
        
    def __params_to_num(self,wb=None):
        for x in self.__params :
            #counter+=1
            # displayer.actualize(10+80*counter/len(self.__params))
            if not(x in ["filtres simples","verif dqc","dqc to osfi","typologies autorisees","specific_months","RT code sit","RT code bat","RT surf"]) :
                if self.__params[x] in [None,""]:
                    self.__params_num[x]=[-1]
                elif isinstance(self.__params[x],list) :
                    self.__params_num[x]=[]
                    for y in self.__params[x] :
                        self.__params_num[x].append(self.get_num_from_col_title(y,wb))
                elif isinstance(self.__params[x],str) :
                    self.__params_num[x]=[self.get_num_from_col_title(self.__params[x],wb)]
                elif isinstance(self.__params[x],int) :
                    self.__params_num[x]=[self.get_num_from_col_title(str(self.__params[x]),wb)]
                elif isinstance(self.__params[x],float) :
                    self.__params_num[x]=[self.get_num_from_col_title(str(self.__params[x]),wb)]
                elif isinstance(self.__params[x],dict) :
                    self.__params_num[x]=[self.get_num_from_col_title(str(self.__params[x]["title"]),wb)]
    def load_excel(self,xls_name):
        loading=True
        if os.path.exists(os.path.join(self.path,xls_name[:-5]+".json")) :
            self.read_json_param(xls_name[:-5]+".json")
        elif os.path.exists(os.path.join(self.path,"all.json")) :
            self.read_json_param("all.json")
        else :
            loading=False
        if loading :
            if "mensualisé" in xls_name :
                for x in self.__params :
                    if isinstance(self.__params[x],dict) and ("annuel" in self.__params[x]) and("mensualisé" in self.__params[x]) :
                        self.__params[x]=self.__params[x]["mensualisé"]
                        
            else :
                for x in self.__params :
                    if isinstance(self.__params[x],dict) and ("annuel" in self.__params[x]) and("mensualisé" in self.__params[x]) :
                        self.__params[x]=self.__params[x]["annuel"]
        if loading and os.path.exists(os.path.join(self.path,"dqc_"+xls_name)) :
            self.add_dqc(os.path.join(self.path,"dqc_"+xls_name))
        if loading and ("mensualisé" in xls_name):
            xls_name=self.mens_to_ann(xls_name)
        if loading and not(xls_name==""):
            # displayer = progression("Lecture du fichier Excel "+xls_name)
            # displayer.actualize(0)
            self.loaded=True
            try :
                self.workbook = openpyxl.load_workbook(os.path.join(self.path,xls_name))
            except :
                self.loaded=False
            # displayer.actualize(10)
            #counter=0
            if self.loaded :
                self.__params_to_num()
            # displayer.actualize(100)
    def __to_int(self,integer):
        if integer==None :
            return 0
        if isinstance(integer,float):
            return int(integer)
        if isinstance(integer,int):
            return integer
        if isinstance(integer,str):
            while integer.startswith("0"):
                integer=integer[1:]
            integer = integer.replace(',','.')
            if integer.startswith('.') :
                integer = "0"+integer
            if integer=="" :
                return 0
            return int(integer)
        return None
    def __to_float(self,integer):
        if integer==None :
            return 0
        if isinstance(integer,float):
            return integer
        if isinstance(integer,int):
            return integer
        if isinstance(integer,str):
            while integer.startswith("0"):
                integer=integer[1:]
            integer = integer.replace(',','.')
            if integer.startswith('.') :
                integer = "0"+integer
            if integer=="" :
                return 0
            return float(integer)
        return 0
    def get_ymd_from_date(self,date,date_format) :
        date = str(date)
        position,i={},0
        position[date_format[0]]=0
        last_date_format=date_format[0]
        while date_format[i]==last_date_format :
            i+=1
        if date_format[i] in ["Y","M","D"] :
            print("You don't have separator inside date format")
            return None
        else :
            sep1 = date_format[i]
            i+=1
            position[date_format[i]]=1
            last_date_format = date_format[i]
            while date_format[i]==last_date_format :
                i+=1
            if date_format[i] in ["Y","M","D"] :
                print("You don't have separator inside date format")
                return None
            else :
                sep2 = date_format[i]
                i+=1
                position[date_format[i]]=1
                if sep1 == sep2 :
                    return self.__to_int(date.split(sep1)[position["Y"]]),self.__to_int(date.split(sep1)[position["M"]]),self.__to_int(date.split(sep1)[position["D"]])
                else :
                    L=[date.split(sep1)[0],date.split(sep1)[1].split(sep2)[0],date.split(sep1)[1].split(sep2)[0]]
                    return self.__to_int(L[position["Y"]]),self.__to_int(L[position["M"]]),self.__to_int(L[position["D"]])
                    
    def mens_to_ann(self,xls_name):
        print("Annuélisation du fichier mensuel\n\tLecture du ficher mensuel")
        loaded=True
        Titles = []
        values=[]
        Values=[]
        Lbat=[]
        Ly=[]
        try :
            wb = openpyxl.load_workbook(os.path.join(self.path,xls_name))
        except :
            print("Error : readinf the mensual excel file")
            loaded=False
        print("\tFichier mensuel Ouvert")
        if loaded :
            self.__params_to_num(wb)
            i=1
            title = wb.active.cell(row=1,column=i).value
            while not(title in [None,"",'""']) :
                Titles.append(title)
                i+=1
                title = wb.active.cell(row=1,column=i).value
            row,flag=2,True
            while flag :
                values.append([])
                for i in range(len(Titles)) :
                    values[-1].append(wb.active.cell(row=row,column=i+1).value)
                flag = False
                for i in range(len(Titles)) :
                    flag = flag or not(wb.active.cell(row=row+1,column=i+1).value in [None,"",'""'])
                    if flag :
                        break
                row+=1
            wb.close()
            print("\tFichier mensuel Lu")
            #Récupération des codes bâtiments et des années
            print("\tConversion des données")
            for i in range(len(values)) :
                year,month,day = self.get_ymd_from_date(values[i][self.__params_num["date"][0]-1],self.__params["date"]["format"])
                bat = values[i][self.__params_num["batiment"][0]-1]
                if not(year in Ly) :
                    Ly.append(year)
                if not(bat in Lbat):
                    Lbat.append(bat)
            for Y in Ly :
                for B in Lbat :
                    val,first = [],True
                    n = len(values)
                    for i in range(n) :
                        year,month,day = self.get_ymd_from_date(values[n-1-i][self.__params_num["date"][0]-1],self.__params["date"]["format"])
                        bat = values[n-1-i][self.__params_num["batiment"][0]-1]
                        if year==Y and bat==B :
                            if first :
                                for j in range(len(values[n-1-i])) :
                                    val.append(values[n-1-i][j])
                                for j in self.__params_num["consommation chaud"] :
                                    val[j-1]=self.__to_float(val[j-1])
                                for j in self.__params_num["consommation froid"] :
                                    val[j-1]=self.__to_float(val[j-1])
                                for j in self.__params_num["consommation non thermique"] :
                                    val[j-1]=self.__to_float(val[j-1])
                                for specific_type in self.__params["specific_months"]:
                                    if month ==self.__params["specific_months"][specific_type]:
                                        for j in self.__params_num["consommation chaud"] :
                                            val.append(values[n-1-i][j-1])
                                        for j in self.__params_num["consommation froid"] :
                                            val.append(values[n-1-i][j-1])
                                        for j in self.__params_num["consommation non thermique"] :
                                            val.append(values[n-1-i][j-1])
                                    else :
                                        for j in self.__params_num["consommation chaud"] :
                                            val.append(0)
                                        for j in self.__params_num["consommation froid"] :
                                            val.append(0)
                                        for j in self.__params_num["consommation non thermique"] :
                                            val.append(0)
                                first = False
                            else :
                                index=0
                                for specific_type in self.__params["specific_months"]:
                                    if month ==self.__params["specific_months"][specific_type]:
                                        for j in self.__params_num["consommation chaud"] :
                                            val[len(Titles)+index]+=self.__to_float(values[n-1-i][j-1])
                                            index+=1
                                        for j in self.__params_num["consommation froid"] :
                                            val[len(Titles)+index]+=self.__to_float(values[n-1-i][j-1])
                                            index+=1
                                        for j in self.__params_num["consommation non thermique"] :
                                            val[len(Titles)+index]+=self.__to_float(values[n-1-i][j-1])
                                            index+=1
                                    else :
                                        for j in self.__params_num["consommation chaud"] :
                                            index+=1
                                        for j in self.__params_num["consommation froid"] :
                                            index+=1
                                        for j in self.__params_num["consommation non thermique"] :
                                            index+=1
                                for j in self.__params_num["consommation chaud"] :
                                    val[j-1]+=self.__to_float(values[n-1-i][j-1])
                                for j in self.__params_num["consommation froid"] :
                                    val[j-1]+=self.__to_float(values[n-1-i][j-1])
                                for j in self.__params_num["consommation non thermique"] :
                                    val[j-1]+=self.__to_float(values[n-1-i][j-1])
                            del values[n-1-i]
                    Values.append(val)
            for specific_type in self.__params["specific_months"]:
                self.__params[specific_type+" consommation chaud"]=[]
                self.__params[specific_type+" consommation froid"]=[]
                self.__params[specific_type+" consommation non thermique"]=[]
                for j in self.__params_num["consommation chaud"] :
                    Titles.append(specific_type+"_"+Titles[j-1])
                    self.__params[specific_type+" consommation chaud"].append(Titles[-1])
                for j in self.__params_num["consommation froid"] :
                    Titles.append(specific_type+"_"+Titles[j-1])
                    self.__params[specific_type+" consommation froid"].append(Titles[-1])
                for j in self.__params_num["consommation non thermique"] :
                    Titles.append(specific_type+"_"+Titles[j-1])
                    self.__params[specific_type+" consommation non thermique"].append(Titles[-1])
            print("\tEnregistrement des données converties")
            wb = openpyxl.Workbook()
            for i in range(len(Titles)):
                wb.active.cell(row=1,column=1+i).value = Titles[i]
                for j in range(len(Values)) :
                    wb.active.cell(row=2+j,column=1+i).value = Values[j][i]
            wb.save(os.path.join(self.path,"annualized_"+xls_name))
            wb.close()
            print("\tDonnées annualisész converties enregistrées")
            return "annualized_"+xls_name
        else :
            return ""
            
        
    def get_num_from_col_title(self,title,wb=None):
        if wb==None :
            if self.loaded :
                count,index,flag=0,0,True
                while count<2 and flag :
                    value = self.workbook.active.cell(row=1,column=index+1).value
                    if value in ['"'+title+'"',title] :
                        return index+1
                    elif value in [None,"",'""']:
                        count+=1
                        index+=1
                    else :
                        count=0
                        index+=1
                return -1
            else :
                return -1
        else :
            count,index,flag=0,0,True
            while count<2 and flag :
                value = wb.active.cell(row=1,column=index+1).value
                if value in ['"'+title+'"',title] :
                    return index+1
                elif value in [None,"",'""']:
                    count+=1
                    index+=1
                else :
                    count=0
                    index+=1
            return -1
    def get_list_of_typologies(self) :
        # displayer = progression("Récupération des différentes typologies de bâtiment")
        # displayer.actualize(0)
        Typologies=[]
        if self.loaded :
            if not(self.__params_num["typologie"][0]==-1) :
                count,index=0,0
                # displayer.actualize((1-1/(index/1000+1))*100)
                while count<100 :
                    value_prev = self.workbook.active.cell(row=index+2,column=self.__params_num["typologie"][0]).value
                    value = self.workbook.active.cell(row=index+2,column=self.__params_num["typologie"][1]).value
                    if (value in [None,"",'""']) and (value_prev in [None,"",'""']):
                        count+=1
                        index+=1
                    elif [value_prev,value] in Typologies :
                        count=0
                        index+=1
                    else :
                        Typologies.append([value_prev,value])
                        count=0
                        index+=1
        # displayer.actualize(100)
        return Typologies
    def __to_val(self,row,col):
        if self.loaded :
            value = self.workbook.active.cell(row=row,column=col).value
            if value in [None,"",'""'] :
                return 0
            if isinstance(value,str) :
                try :
                    v=float(value.replace("'",'').replace('"','').replace(",","."))
                except :
                    return 0
            else :
                if isinstance(value,int) or isinstance(value,float) :
                    v=value
            return v
        return 0
    def get_values_for_typologie(self,typologie):
        print("Travaille sur la typologie "+str(typologie))
        # displayer = progression("Récupération des grandeurs pour la typologie "+typologie)
        # displayer.actualize(0)
        L_values=[]
        count,index=0,0
        while count<100 :
            # displayer.actualize((1-1/(index/1000+1))*100)
            value_prev = self.workbook.active.cell(row=index+2,column=self.__params_num["typologie"][0]).value
            value = self.workbook.active.cell(row=index+2,column=self.__params_num["typologie"][1]).value
            if (value in [None,"",'""']) and (value_prev in [None,"",'""']):
                count+=1
                index+=1
            elif [value_prev,value] == typologie :
                if self.__params["filtres simples"]==None :
                    for x in self.to_remove_dqc :
                        flag = False
                        for attr in x :
                            attr_index = self.get_num_from_col_title(attr)
                            flag = flag or  (attr_index==-1) or not(self.workbook.active.cell(row=index+2,column=attr_index).value==x[attr])
                            if flag :
                                break
                        if not(flag):
                            break
                    if flag :
                        L_values.append({})
                        for elem in self.__params_num :
                            if not(elem=="typologie") :
                                L_values[-1][elem]=0
                                for new_index in self.__params_num[elem] :
                                    if (new_index>=0) :
                                        have_to_be_num = False
                                        for start in self.__params["specific_months"] :
                                            if elem.startswith(start+" "):
                                                have_to_be_num = True
                                                break
                                        if have_to_be_num or elem.startswith("Surface") or elem.startswith("DJ") or elem.startswith("consommation") :
                                            L_values[-1][elem]+=self.__to_val(row=index+2,col=new_index)
                                        else :
                                            L_values[-1][elem]=self.workbook.active.cell(row=index+2,column=new_index).value
                        L_values[-1]["typologie"]=typologie
                        L_values[-1]["dqc error"]=False
                    else :
                        L_values.append({})
                        for elem in self.__params_num :
                            if not(elem=="typologie") :
                                L_values[-1][elem]=0
                                for new_index in self.__params_num[elem] :
                                    if (new_index>=0) :
                                        have_to_be_num = False
                                        for start in self.__params["specific_months"] :
                                            if elem.startswith(start+" "):
                                                have_to_be_num = True
                                                break
                                        if have_to_be_num or elem.startswith("Surface") or elem.startswith("DJ") or elem.startswith("consommation") :
                                            L_values[-1][elem]+=self.__to_val(row=index+2,col=new_index)
                                        else :
                                            L_values[-1][elem]=self.workbook.active.cell(row=index+2,column=new_index).value
                        L_values[-1]["typologie"]=typologie
                        L_values[-1]["dqc error"]=True
                else :
                    flag=True
                    for Filter in self.__params["filtres simples"] :
                        filter_index = self.get_num_from_col_title(Filter)
                        flag = flag and ((filter_index==-1) or (self.workbook.active.cell(row=index+2,column=filter_index).value==self.__params["filtres simples"][Filter]))
                        if not(flag):
                            break
                    if flag :
                        for x in self.to_remove_dqc :
                            flag = False
                            for attr in x :
                                attr_index = self.get_num_from_col_title(attr)
                                flag = flag or  (attr_index==-1) or not(self.workbook.active.cell(row=index+2,column=attr_index).value==x[attr])
                                if flag :
                                    break
                            if not(flag):
                                break
                        if flag :
                            L_values.append({})
                            for elem in self.__params_num :
                                if not(elem=="typologie") :
                                    L_values[-1][elem]=0
                                    for new_index in self.__params_num[elem] :
                                        if (new_index>=0) :
                                            have_to_be_num = False
                                            for start in self.__params["specific_months"] :
                                                if elem.startswith(start+" "):
                                                    have_to_be_num = True
                                                    break
                                            if have_to_be_num or elem.startswith("Surface") or elem.startswith("DJ") or elem.startswith("consommation") :
                                                L_values[-1][elem]+=self.__to_val(row=index+2,col=new_index)
                                            else :
                                                L_values[-1][elem]=self.workbook.active.cell(row=index+2,column=new_index).value
                            L_values[-1]["typologie"]=typologie
                            L_values[-1]["dqc error"]=False
                        else :
                            L_values.append({})
                            for elem in self.__params_num :
                                if not(elem=="typologie") :
                                    L_values[-1][elem]=0
                                    for new_index in self.__params_num[elem] :
                                        if (new_index>=0) :
                                            have_to_be_num = False
                                            for start in self.__params["specific_months"] :
                                                if elem.startswith(start+" "):
                                                    have_to_be_num = True
                                                    break
                                            if have_to_be_num or elem.startswith("Surface") or elem.startswith("DJ") or elem.startswith("consommation") :
                                                L_values[-1][elem]+=self.__to_val(row=index+2,col=new_index)
                                            else :
                                                L_values[-1][elem]=self.workbook.active.cell(row=index+2,column=new_index).value
                            L_values[-1]["typologie"]=typologie
                            L_values[-1]["dqc error"]=True
                count=0
                index+=1
            else :
                count=0
                index+=1
        # displayer.actualize(100)
        if self.__params["typologies autorisees"]==None :
            return L_values,0
        if ([typologie[0]] in self.__params["typologies autorisees"]) or (typologie in self.__params["typologies autorisees"]):
            return L_values,0
        return [],len(L_values)
        #format de la régression : Conso = A0 + A1*DJU*(S**0.5) + A2*DJF*(S**0.5) +A3*S
    def exclude_neg_surf(self,Lval):
        print("\tExclusion des surfaces négatives (ou nulles)")
        # displayer = progression("Suppression des bâtiments ayant une surface négative")
        # displayer.actualize(0)
        count=0
        n=len(Lval)
        for i in range(n) :
            # displayer.actualize((i+1)/n*100)
            if Lval[n-1-i]["Surface"]<=0  :
                count+=1
                del Lval[n-1-i]
        # displayer.actualize(100)
        print("\t\t"+str(count)+" bâtiments exclus")
        return count
    def exclude_neg_elec(self,Lval):
        print("\tExclusion des consommations électriques négatives (ou nulles)")
        # displayer = progression("Suppression des bâtiments ayant une consomation électrique négative")
        # displayer.actualize(0)
        count=0
        n=len(Lval)
        for i in range(n) :
            # displayer.actualize((i+1)/n*100)
            if Lval[n-1-i]["consommation non thermique"]<=0  :
                count+=1
                del Lval[n-1-i]
        # displayer.actualize(100)
        print("\t\t"+str(count)+" bâtiments exclus")
        return count
    def exclude_strict_neg_consos(self,Lval):
        print("\tExclusion des consommations totales négatives")
        # displayer = progression("Suppression des bâtiments ayant une consomation strictement négative")
        # displayer.actualize(0)
        count=0
        n=len(Lval)
        for i in range(n) :
            # displayer.actualize((i+1)/n*100)
            if Lval[n-1-i]["consommation froid"]<0 or Lval[n-1-i]["consommation chaud"]<0  :
                count+=1
                del Lval[n-1-i]
        # displayer.actualize(100)
        print("\t\t"+str(count)+" bâtiments exclus")
        return count
    def exclude_dqc(self,Lval) :
        print("\tExclusion des erreures relevées dans le DQC si disponibles")
        count=0
        n=len(Lval)
        for i in range(n) :
            if Lval[n-1-i]["dqc error"] :
                count+=1
                del Lval[n-1-i]
        print("\t\t"+str(count)+" bâtiments exclus")
        return count
    def exclude_bad_rt_surf(self,Lval,xls_name):
        print("\tExclusion des surfaces discordantes avec celles relevées dans le RT")
        def get_ending_code(code) :
            if isinstance(code,float):
                return int(code)
            if isinstance(code,int):
                return code
            if isinstance(code,str):
                while code.endswith(" ") or  code.endswith(" ") :
                    code = code[:-1]
                txt=""
                n=len(code)
                for i in range(n):
                    if code[n-1-i] in "0123456789" :
                        txt=code[n-1-i]+txt
                    else :
                        return int(txt)
                return int(txt)
            return None
        def get_starting_code(code):
            if isinstance(code,float):
                return int(code)
            if isinstance(code,int):
                return code
            if isinstance(code,str):
                while code.startswith(" ") or  code.startswith(" ") :
                    code = code[1:]
                txt=""
                for i in range(len(code)):
                    if code[i] in "0123456789" :
                        txt+=code[i]
                    else :
                        if txt=="" :
                            return None
                        return int(txt)
                return int(txt)
            return None
        def compear_code(s1,b1,s2,b2):
            return (str(s1)==str(s2)) and (str(b1)==str(b2))
        def compear_surf(surf_,L_surf_):
            print("\t\tCompearing OSFI surf to RT surf\n\t\t\t"+str(surf_)+" : "+str(L_surf_))
            flag=True
            for i in range(len(L_surf_)):
                flag = flag and L_surf_[i] in [None,"",'""']
                if not(flag):
                    break
            if flag :
                return True
            else :
                for i in range(len(L_surf_)):
                    if not(self.__to_float(L_surf_[i])==0) :
                        if ((surf_-self.__to_float(L_surf_[i]))>=-0.05 and (surf_-self.__to_float(L_surf_[i]))<=0.05) :
                            return True
                return False
        LS=[]
        L_index=[]
        wb = openpyxl.load_workbook(os.path.join(self.path,xls_name))
        L_index.append(self.get_num_from_col_title(self.__params["RT code sit"],wb))
        L_index.append(self.get_num_from_col_title(self.__params["RT code bat"],wb))
        for x in self.__params["RT surf"] :
            L_index.append(self.get_num_from_col_title(x,wb))
        print(L_index)
        i=3
        val = []
        for index in L_index :
            val.append(wb.active.cell(row=i,column=index).value)
        while not(val[0] in ["",None,'""']) :
            i+=1
            LS.append(val)
            val = []
            for index in L_index :
                val.append(wb.active.cell(row=i,column=index).value)
        n = len(Lval)
        L_bat_sit_rt=[]
        for i in range(n) :
            print(Lval[i])
            if not([Lval[i]["code sit RT"],Lval[i]["code bat RT"]] in L_bat_sit_rt) :
                L_bat_sit_rt.append([Lval[i]["code sit RT"],Lval[i]["code bat RT"]])
        print(L_bat_sit_rt)
        L_surf_to_test=[]
        n = len(Lval)
        for [site,bat] in L_bat_sit_rt :
            L_surf_to_test.append(0)
            OSFI_bat=[]
            for i in range(n) :
                if (Lval[i]["code sit RT"]==site) and (Lval[i]["code bat RT"]==bat) and not(Lval[i]["batiment"] in OSFI_bat):
                    OSFI_bat.append(Lval[i]["batiment"])
                    L_surf_to_test[-1]+=Lval[i]["Surface"]
        L_surf_validate = []
        for i in range(len(L_bat_sit_rt)) :
            [site,bat] = L_bat_sit_rt[i]
            bat = get_starting_code(bat)
            site = get_ending_code(site)
            unfounded = True 
            for j in range(len(LS)) :
                if compear_code(site,bat,LS[j][0],LS[j][1]) :
                    L_surf_validate.append(compear_surf(L_surf_to_test[i],LS[j][2:]))
                    unfounded = False
                    break
            if unfounded :
                print("Bat "+str([site,bat])+"not found in RT")
                L_surf_validate.append(True)
        count=0
        for i in range(n) :
            site,bat =Lval[n-1-i]["code sit RT"],Lval[n-1-i]["code bat RT"]
            j = L_bat_sit_rt.index([site,bat])
            if not(L_surf_validate[j]) :
                del Lval[n-1-i]
                count+=1
        print("\t\t"+str(count)+" bâtiments exclus")
        return count
    def exclude_not_heat(self,Lval) :
        print("\tExclusion des consommations sans chauffage (consommation totale de janvier moins de 10% supérieure à la consommation de mai)")
        test=True
        for specific_type in self.__params["specific_months"]:
            test = test and ((specific_type+" consommation chaud") in self.__params) and ((specific_type+" consommation froid") in self.__params) and ((specific_type+" consommation non thermique") in self.__params)
        if test :
            count=0
            specific_type1,specific_type2="rien","chauffage"
            n=len(Lval)
            for i in range(n):
                if (Lval[n-1-i][specific_type1+" consommation chaud"]+Lval[n-1-i][specific_type1+" consommation froid"]+Lval[n-1-i][specific_type1+" consommation non thermique"])*1.1>(Lval[n-1-i][specific_type2+" consommation chaud"]+Lval[n-1-i][specific_type2+" consommation froid"]+Lval[n-1-i][specific_type2+" consommation non thermique"]) :
                    del Lval[n-1-i]
                    count+=1
            print("\t\t"+str(count)+" bâtiments exclus")
            return count
        else :
            print("\t\tAucun bâtiments exclus Faute de résultats initiaux mensualisés")
            return 0
    def exclude_not_cool(self,Lval):
        print("\tExclusion des consommations sans clim (consommation totale de juillet moins de 10% supérieure à la consommation de mai)")
        test=True
        for specific_type in self.__params["specific_months"]:
            test = test and ((specific_type+" consommation chaud") in self.__params) and ((specific_type+" consommation froid") in self.__params) and ((specific_type+" consommation non thermique") in self.__params)
        if test :
            count=0
            specific_type1,specific_type2="rien","clim"
            n=len(Lval)
            for i in range(n):
                if (Lval[n-1-i][specific_type1+" consommation chaud"]+Lval[n-1-i][specific_type1+" consommation froid"]+Lval[n-1-i][specific_type1+" consommation non thermique"])*1.1>(Lval[n-1-i][specific_type2+" consommation chaud"]+Lval[n-1-i][specific_type2+" consommation froid"]+Lval[n-1-i][specific_type2+" consommation non thermique"]) :
                    del Lval[n-1-i]
                    count+=1
            print("\t\t"+str(count)+" bâtiments exclus")
            return count
            
        else :
            print("\t\tAucun bâtiments exclus Faute de résultats initiaux mensualisés")
            return 0
    def create_meta_excel(self,excel_name,rt_doc,alpha=0.05):
        LL=[["typologie","Typologie détaillée","Nombre de Bâtiments initiaux",
             "Bâtiments supprimés","","","","","","","Nombre de bâtiments utilisables",
             "Régression par rapport à la surface","","","","","","",
             "Régression par rapport à la surface et aux DJU","","","","","","","","","",
             ],
            ["Typologie","Typologie détaillée","Nombre de Bâtiments initiaux",
             "Nombre de bâtiments exclus car hors périmètre",
             "Nombre de bâtiments exclus pour surface négative",
             "Nombre de bâtiments exclus pour consommation électrique négative",
             "Nombre de bâtiments exclus pour une consommation strictement négative",
             "Nombre de bâtiments exclus par l'annalyse DQC",
             "Nombre de bâtiments exclue pour surface incohérente avec celle du RT",
             "Nombre de bâtiments exclue pour une trop faible augmentation des consommations en janvier",
             "Nombre de bâtiments utilisables",
             "r²",
             "Offset de la consommation","Pente de la consommation par rapport à la surface",
             "Offset de la consommation minimale","Pente de la consommation par rapport à la surface minimale",
             "Offset de la consommation maximale","Pente de la consommation par rapport à la surface maximale",
             "r²",
             "Offset de la consommation","Pente de la consommation par rapport à la surface","Pente de la consommation par rapport aux DJU multiplié par la racine carrée de la surface",
             "Offset de la consommation minimale","Pente de la consommation par rapport à la surface minimale","Pente de la consommation par rapport aux DJU multiplié par la racine carrée de la surface minimale",
             "Offset de la consommation maximale","Pente de la consommation par rapport à la surface maximale","Pente de la consommation par rapport aux DJU multiplié par la racine carrée de la surface maximale"]]
        self.load_excel(excel_name)
        List_of_typologies = self.get_list_of_typologies()
        for x in List_of_typologies :
            l = [x[0],x[1]]
            L,rem=self.get_values_for_typologie(x)
            l.append(len(L)+rem)
            if len(L)>0 :
                l.append(rem)
                l.append(self.exclude_neg_surf(L))
                l.append(self.exclude_neg_elec(L))
                l.append(self.exclude_strict_neg_consos(L))
                l.append(self.exclude_bad_rt_surf(L,rt_doc))
                l.append(self.exclude_not_heat(L))
                l.append(self.exclude_dqc(L))
                l.append(len(L))
                X1,X2,Y=[],[],[]
                for y in L :
                    Y.append(y['consommation chaud']+y['consommation froid']+y['consommation non thermique'])
                    X1.append([]),X2.append([])
                    X1[-1].append(1)
                    X2[-1].append(1)
                    X1[-1].append(y["Surface"])
                    X2[-1].append(y["Surface"])
                    X2[-1].append((y["Surface"]**0.5)*(y["DJU"]))
                if (len(Y)>2) :
                    regression1=sm.OLS(Y,X1).fit()
                    l.append(regression1.rsquared)
                    params=regression1.params
                    Iconf=regression1.conf_int(alpha=alpha)
                    l.append(params[0]),l.append(params[1])
                    l.append(Iconf[0][1]),l.append(Iconf[1][1])
                    l.append(Iconf[0][0]),l.append(Iconf[1][0])
                    regression2=sm.OLS(Y,X2).fit()
                    l.append(regression2.rsquared)
                    params=regression2.params
                    Iconf=regression2.conf_int(alpha=alpha)
                    l.append(params[0]),l.append(params[1]),l.append(params[2])
                    l.append(Iconf[0][1]),l.append(Iconf[1][1]),l.append(Iconf[2][1])
                    l.append(Iconf[0][0]),l.append(Iconf[1][0]),l.append(Iconf[2][0])
                else :
                    l.append("Contient moins de 2 batiments")
            else :
                l.append("hors du périmètre d'étude")
            LL.append(l)
        self.loaded=False
        try :
            self.workbook.close()
        except :
            self.loaded=True
        workbook = openpyxl.Workbook()
        for line in range(len(LL)):
            for col in range(len(LL[line])) :
                if (LL[line][col]==None) or isinstance(LL[line][col],int) or isinstance(LL[line][col],float) :
                    workbook.active.cell(row=line+1,column=col+1).value = LL[line][col]
                else :
                    workbook.active.cell(row=line+1,column=col+1).value = str(LL[line][col])
        if (self.dqc_done) :
            excel_name = "dqc_"+excel_name
        workbook.save(os.path.join(self.path,"meta_"+excel_name))
        workbook.close()
if __name__=="__main__" :
    path = os.path.join(os.path.dirname(os.getcwd()),"Datas")
    L_files = os.listdir(path)
    rt_doc = "Surfaces_OAD_2024_08_05.xlsx"
    L_files=["Consommations mensualisées des équipements_simplifie.xlsx"]
    for file in L_files :
        XL = OSFI_regress(path)
        if not(file.startswith("meta_") or file.startswith("dqc_") or file.startswith("Surface_OAD")) and file.endswith(".xlsx") :
            XL.create_meta_excel(file,rt_doc)
        del XL
            
        
        
