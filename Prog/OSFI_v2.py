# -*- coding: utf-8 -*-
"""
Created on Tue Aug 20 15:21:31 2024

@author: sacha.mailler
"""
import openpyxl
from Avancement import progression
class Excel:
    def __init__(self,path=None) :
        self.Names=[]
        self.Valuse=[]
        self.flag=False
        self.path=path
        self.aux_loaded=""
        self.aux_workbook=None
        if not(path==None) :
            self.read()
    def __is(self,a,b):
        if isinstance(a,list) and not(isinstance(b,list)):
            for x in a :
                if self.__is(x,b):
                    return True
            return False
        if isinstance(b,list) and not(isinstance(a,list)):
            for x in b :
                if self.__is(a,x):
                    return True
            return False
        if a==None :
            return b in [None,"",'""']
        if isinstance(a,float):
            _a_=str(a)
        elif isinstance(a,int):
            _a_=str(a)
        elif isinstance(a,str):
            _a_=a
            while _a_.startswith('"') and _a_.endswith('"'):
                _a_=_a_[1:-1]
        if isinstance(b,float):
            _b_=str(b)
        elif isinstance(b,int):
            _b_=str(b)
        elif isinstance(b,str):
            _b_=b
            while _b_.startswith('"') and _b_.endswith('"'):
                _b_=_b_[1:-1]
        elif b==None :
            return a in [None,"",'""']
        else :
            print("Erreur de comparaison : ")
            print("\t"+str(a))
            print("\t"+str(b))
        return _a_==_b_
    def __to_num(self,a):
        if a==None :
            return 0
        if isinstance(a,float) or isinstance(a,int) :
            return a
        if isinstance(a,str):
            val=-1
            _a_=a
            while _a_.startswith('"') and _a_.endswith('"') :
                _a_=_a_[1:-1]
            if _a_=="" :
                return 0
            try :
                val=float(_a_)
            except :
                try :
                    val = float(_a_.replace(',','.'))
                except :
                    print("Erreur de convertion : "+a+" n'est pas un nombre")
            return val
        print("Erreur de convertion : "+a+" n'est pas un nombre")
        return -1
    def __get_sig(self,value,known_element={},m=None):
        main_index = self.index(value)
        Indexes=[]
        Values=[]
        variance=0
        nv=0
        for x in known_element :
            Indexes.append(self.index(x))
            Values.append(known_element[x])
        if m==None :
            m=self.__get_mean()
        for i in range(len(self.Values)) :
            flag = True
            for j in range(len(Indexes)) :
                flag = self.__is(self.Values[i][Indexes[j]],Values[j])
                if not(flag) :
                    break
            if flag :
                if not(self.__to_num(self.Values[i][main_index])==0) :
                    nv=nv+1
                    variance+=(m-self.__to_num(self.Values[i][main_index]))*(m-self.__to_num(self.Values[i][main_index]))
        variance/=nv
        return variance**0.5
    def __get_mean(self,value,known_element={}):
       main_index = self.index(value)
       Indexes=[]
       Values=[]
       nm=0
       for x in known_element :
           Indexes.append(self.index(x))
           Values.append(known_element[x])
       m=0
       for i in range(len(self.Values)) :
           flag = True
           for j in range(len(Indexes)) :
               flag = self.__is(self.Values[i][Indexes[j]],Values[j])
               if not(flag) :
                   break
           if flag :
               if not(self.__to_num(self.Values[i][main_index])==0) :
                   nm=nm+1
                   m+=(self.__to_num(self.Values[i][main_index]))
       m/=nm
       return m
    def read(self,path=None) :
        self.Names=[]
        self.Values=[]
        self.flag=False
        if path==None :
            PR = progression("Lecture du fichier :"+self.path.split("\\")[-1].split('/')[-1])
            if self.path==None :
                self.flag = False
            else :
                workbook = openpyxl.load_workbook(self.path)
        else :
            PR = progression("Lecture du fichier :"+path.split("\\")[-1].split('/')[-1])
            workbook = openpyxl.load_workbook(path)
        n=workbook.active.max_column
        for col in range(n):
            PR.actualize(5*(col+1)/n)
            self.Names.append(workbook.active.cell(row=1,column=col+1).value)
        n=workbook.active.max_row-1
        for rw in range(n):
            PR.actualize(5+94*(rw+1)/n)
            self.Values.append([])
            for col in range(workbook.active.max_column):
                self.Values[-1].append(workbook.active.cell(row=rw+2,column=col+1).value)
        workbook.close()
        PR.actualize(100)
    def index(self,name):
        if isinstance(name,str):
            flag=False
            for i in range(len(self.Names)):
                if self.__is(self.Names[i],name):
                    flag=True
                    break
            if flag :
                return i
            return -1
        print("Erreur de repérage : le nom d'une colonne doit être une chaîne de caractères")
    def add_values(self,known_element,element_to_add):
        for x in element_to_add :
            if self.index(x)==-1 :
                self.Names.append(x)
                for i in range(len(self.Values)):
                    self.Values[i].append(None)
        I=[]
        I2=[]
        for x in known_element :
            i=self.index(x)
            if i==-1 :
                print("Erreur d'attribution d'une valeur : l'un des éléments à reconnaitre n'est pas dans le fichier de base")
            else :
                I.append(i)
                I2.append(x)
        for i in range(len(self.Values)):
            flag=False
            for j in range(len(I)):
                flag = not(self.__is(self.Values[i][I[j]],known_element[I2[j]]))
                if flag :
                    break
            if not(flag) :
                for x in element_to_add :
                    self.Values[i][self.index(x)] = element_to_add[x]
                    
    def is_in(self,col_names,col_values,final_name):
        self.Names.append(final_name)
        for i in range(len(self.Values)) :
            for j in range(len(col_values)):
                flag=True
                for k in range(len(col_values[j])):
                    flag = flag and self.__is(self.Values[i][self.index(col_names[k])],col_values[j][k])
                if flag :
                    break
            self.Values[i].append(flag)
                
    def sum_values(self,known_element,element_to_add):
        for x in element_to_add :
            if self.index(x)==-1 :
                self.Names.append(x)
                for i in range(len(self.Values)):
                    self.Values[i].append(None)
        I=[]
        I2=[]
        for x in known_element :
            i=self.index(x)==-1 
            if i==-1 :
                print("Erreur d'attribution d'une valeur : l'un des éléments à reconnaitre n'est pas dans le fichier de base")
            else :
                I.append(i)
                I2.append(x)
        for i in range(len(self.Values)):
            flag=False
            for j in range(I):
                flag = flag or not(self.__is(self.Values[i][I[j]],known_element[I2[i]]))
                if flag :
                    break
            if not(flag) :
                for x in element_to_add :
                    self.Values[i][self.index(x)] = self.__to_num(self.Values[i][self.index(x)])+self.__to_num(element_to_add[x])
    def virtual_group_by_sum(self,known_element,to_sum_name,result_prefix="group_",add_count=False) :
        # PR = progression("Ajout des colonne "+str(to_sum_name)+" de la somme groupée sur "+str(known_element))
        concerned_lines=[]
        I=[]
        I2=[]
        for x in known_element :
            i=self.index(x)
            if i==-1 :
                print("Erreur d'attribution d'une valeur : l'un des éléments à reconnaitre n'est pas dans le fichier de base")
            else :
                I.append(i)
                I2.append(x)
        # PR.actualize(0)
        n=len(self.Values)
        for i in range(len(self.Values)):
            flag=False
            for j in range(len(I)):
                flag = flag or not(self.__is(self.Values[i][I[j]],known_element[I2[j]]))
                if flag :
                    break
            if not(flag) :
                concerned_lines.append(i)
        if add_count :
            add_count_index = self.index(result_prefix+"count")
            if add_count_index==-1 :
                add_count_index=len(self.Names)
                self.Names.append(result_prefix+"count")
                for i in range(len(self.Values)) :
                    self.Values[i].append(None)
        count = len(concerned_lines)
        init_indexes=[]
        final_indexes=[]
        for j in range(len(to_sum_name)) :
            i = self.index(to_sum_name[j])
            if not(i==-1) :
                init_indexes.append(i)
                i2 = self.index(result_prefix+to_sum_name[j])
                if i2==-1 :
                    final_indexes.append(len(self.Names))
                    self.Names.append(result_prefix+to_sum_name[j])
                    for i3 in range(len(self.Values)) :
                        self.Values[i3].append(None)
                else :
                    final_indexes.append(i2)
        for i in range(count) :
            if add_count :
                self.Values[concerned_lines[i]][add_count_index]=count
        for j in range(len(init_indexes)) :
            value = 0
            for i in range(count) :
                value+=self.__to_num(self.Values[concerned_lines[i]][init_indexes[j]])
            for i in range(count) :
                self.Values[concerned_lines[i]][final_indexes[j]]=value
    def import_columns_from(self,path,links_main_to_imported,cols_to_import,where={},collapsed="summ",count_imported=False) :
        if count_imported :
            count_index = len(self.Names)
            name = "Nombre d éléments importés"
            if self.index(name)==-1 :
                self.Names.append(name)
            else :
                endname=1
                while not(self.index(name+" "+str(endname))==-1) :
                    endname+=1
                self.Names.append(name+" "+str(endname))
            for i in range(len(self.Values)) :
                self.Values[i].append(0)
        L_names=[]
        L_values=[]
        def index(name):
            for i in range(len(L_names)) :
                if self.__is(name,L_names[i]) :
                    return i
            return -1
        PR = progression("Lecture du fichier :"+path.split("\\")[-1].split('/')[-1])
        if self.aux_loaded=="" or not(self.aux_loaded==path) :
            if not(self.aux_loaded=="") :
                self.aux_workbook.close()
            self.aux_loaded=path
            self.aux_workbook = openpyxl.load_workbook(path)
        PR.actualize(1)
        n=self.aux_workbook.active.max_column
        for col in range(self.aux_workbook.active.max_column):
            PR.actualize(1+4*(col+1)/n)
            L_names.append(self.aux_workbook.active.cell(row=1,column=col+1).value)
        n=self.aux_workbook.active.max_row-1
        for rw in range(self.aux_workbook.active.max_row-1):
            PR.actualize(5+94*(rw+1)/n)
            L_values.append([])
            for col in range(self.aux_workbook.active.max_column):
                L_values[-1].append(self.aux_workbook.active.cell(row=rw+2,column=col+1).value)
        PR.actualize(100)
        PR = progression("Filtrage")
        PR.actualize(0)
        ncount,count=len(where),1
        for x in where :
            PR.actualize(count/ncount)
            i = index(x)
            if not(i==-1) :
                n=len(L_values)
                for j in range(n) :
                    if not(self.__is(L_values[n-1-j][i],where[x])) :
                        del L_values[n-1-j]
        PR.actualize(100)
        L_import_index = []
        L_imported_indexes=[]
        if isinstance(cols_to_import,list):
            for x in cols_to_import :
                i = index(x)
                if not(i==-1) :
                    L_import_index.append(i)
                    i = self.index(x)
                    if i==-1 :
                        L_imported_indexes.append(len(self.Names))
                        self.Names.append(x)
                        for i in range(len(self.Values)) :
                            self.Values[i].append(None)
                    else :
                        L_imported_indexes.append(i)
                        print("Attention dans l'import des colonnes vous impoprtez des colonnes existant déjà")
        elif isinstance(cols_to_import,dict) :
            for x in cols_to_import :
                i = index(x)
                if not(i==-1) :
                    L_import_index.append(i)
                    i = self.index(cols_to_import[x])
                    if i==-1 :
                        L_imported_indexes.append(len(self.Names))
                        self.Names.append(cols_to_import[x])
                        for i in range(len(self.Values)) :
                            self.Values[i].append(None)
                    else :
                        L_imported_indexes.append(i)
                        print("Attention dans l'import des colonnes vous impoprtez des colonnes existant déjà")
        links_base=[]
        links_imported=[]
        for x in  links_main_to_imported :
            i,j=self.index(x),index(links_main_to_imported[x])
            if not(i==-1) and not(j==-1) :
                links_base.append(i)
                links_imported.append(j)
        n = len(self.Values)
        PR = progression("Ajout des éléments de la table chargée (partie longue de l'algorithme)")
        for j in range(len(self.Values)) :
            PR.actualize(99*(j+1)/n)
            for i in range(len(L_values)) :
                flag=True
                for k in range(len(links_base)):
                    flag = self.__is(self.Values[j][links_base[k]],L_values[i][links_imported[k]])
                    if not(flag) :
                        break
                if flag :
                    if(count_imported) :
                        self.Values[j][count_index] = self.__to_num(self.Values[j][count_index])+1
                    for k in range(len(L_import_index)) :
                        if collapsed=="summ" :
                            self.Values[j][L_imported_indexes[k]] = self.__to_num(self.Values[j][L_imported_indexes[k]])+self.__to_num(L_values[i][L_import_index[k]])
                        elif collapsed=="concat" :
                            if self.Values[j][L_imported_indexes[k]] in [None,"",'""'] :
                                self.Values[j][L_imported_indexes[k]]=str(L_values[i][L_import_index[k]])
                            else :
                                self.Values[j][L_imported_indexes[k]]=str(self.Values[j][L_imported_indexes[k]])+";"+str(L_values[i][L_import_index[k]])
                        else :
                            self.Values[j][L_imported_indexes[k]]=L_values[i][L_import_index[k]]
        PR.actualize(100)
    def save(self,path):
        if path.endswith(".xlsx") :
            PR = progression("Enregistrement du fichier : "+path.split("\\")[-1].split("/")[-1])
            workbook = openpyxl.Workbook()
            PR.actualize(0)
            n=len(self.Names)
            for i in range(n) :
                PR.actualize(100*(i+1)/n)
                workbook.active.cell(row=1,column = i+1).value = self.Names[i]
                for j in range(len(self.Values)) :
                    workbook.active.cell(row=j+2,column = i+1).value = self.Values[j][i]
            PR.actualize(100)
            workbook.save(path)
            workbook.close()
        elif path.endswith(".csv") :
            PR = progression("Enregistrement du fichier : "+path.split("\\")[-1].split("/")[-1])
            PR.actualize(0)
            with open(path,"w",encoding="utf-8") as file :
                n=len(self.Names)
                for i in range(len(self.Names)) :
                    PR.actualize(5*(i+1)/n)
                    file.write(self.Names[i].replace(";",",")+";")
                file.write("\n")
                n = len(self.Values)
                for i in range(n) :
                    PR.actualize(5+94*(i+1)/n)
                    for j in range(len(self.Values[i])) :
                        file.write(str(self.Values[i][j]).replace(";",",")+";")
                    file.write("\n")
                PR.actualize(100)
        else :
            print("Le type de fichier pour l'enregistrement est inconnu")
    def get_list_from_col(self,name,known_element={}):
        PR = progression("Récupération des valeurs possibles de la colonne : "+name)
        PR.actualize(0)
        i = self.index(name)
        if i==-1 :
            return []
        Indexes = []
        Verifs = []
        for x in known_element :
            I = self.index(x)
            if not(I==-1) :
                Indexes.append(I)
                Verifs.append(known_element[x])
            
        L=[]
        n = len(self.Values)
        for j in range(n) :
            PR.actualize((j+1)/n*100)
            flag=True
            for I in range(len(Indexes)) :
                if not(self.__is(self.Values[j][Indexes[I]],Verifs[I])) :
                    flag=False
                    break
            if flag and not(self.Values[j][i] in L) :
                L.append(self.Values[j][i])
        PR.actualize(100)
        return L
    def get_list_from_cols(self,names,known_element={}):
        PR = progression("Récupération des valeurs possibles des colonnes : "+str(names))
        PR.actualize(0)
        inames =[]
        for name in names :
            inames.append(self.index(name))
        Indexes = []
        Verifs = []
        for x in known_element :
            I = self.index(x)
            if not(I==-1) :
                Indexes.append(I)
                Verifs.append(known_element[x])
            
        L=[]
        n = len(self.Values)
        for j in range(n) :
            PR.actualize((j+1)/n*100)
            flag=True
            for I in range(len(Indexes)) :
                if not(self.__is(self.Values[j][Indexes[I]],Verifs[I])) :
                    flag=False
                    break
            if flag :
                l=[]
                for i in inames :
                    if i==-1 :
                        l.append(None)
                    else :
                        l.append(self.Values[j][i])
            if flag and not(l in L) :
                L.append(l)
        PR.actualize(100)
        return L
    def remove(self,known_element):
        Indexes=[]
        Values=[]
        for x in known_element :
            i=self.index(x)
            if not(i==-1) :
                Indexes.append(i)
                Values.append(known_element[x])
        n=len(self.Values)
        for i in range(n) :
            flag=True
            for j in range(len(Indexes)) :
                if not(self.__is(self.Values[n-1-i][Indexes[j]],Values[j])) :
                    flag=False
                    break
            if flag :
                del self.Values[n-1-i]
    def create_concat_col(self,L_names,name=None):
        Indexes=[]
        if name==None :
            Name=""
        else :
            Name=name
        for x in L_names :
            i = self.index(x)
            if not(i==-1) :
                if name==None :
                    Name+=x
                Indexes.append(i)
        self.Names.append(Name)
        for i in range(len(self.Values)) :
            var =""
            for j in Indexes :
                if not(self.Values[i][j]==None) :
                    var+=str(self.Values[i][j])
            self.Values[i].append(var)
    def __extract_ending_num(self,value):
        if isinstance(value,int):
            return str(value)
        if isinstance(value,float):
            return str(value)
        if isinstance(value,str):
            while value.endswith(" ") :
                value=value[:-1]
            val=""
            n=len(value)
            cp=0
            for i in range(n):
                if value[n-1-i] in [".",","] :
                    if cp==0 :
                        cp=1
                        val=value[n-1-i]+val
                    else :
                        break
                elif value[n-1-i] in "0123456789" :
                    val=value[n-1-i]+val
                else :
                    break
            return val
        return ""
    def extract_ending_num(self,columns):
        Indexes=[]
        Indexes_o=[]
        for x in columns :
            i=self.index(x)
            if not(i==-1) :
                Indexes.append(i)
                i=self.index(columns[x])
                if (i==-1) :
                    Indexes_o.append(len(self.Names))
                    self.Names.append(columns[x])
                    for j in range(len(self.Values)) :
                        self.Values[j].append(None)
                else :
                    Indexes_o.append(i)
        for i in range(len(self.Values)) :
            for j in range(len(Indexes)) :
                self.Values[i][Indexes_o[j]] = self.__extract_ending_num(self.Values[i][Indexes[j]])
    def get_stat_by_element(self,col,value,known_element={}):
        L = XL.get_list_from_col(col,known_element)
        i=0
        for x in L :
            print(i)
            i+=1
            known_element[col] = x
            m = self.__get_mean(value,known_element)
            sig = self.__get_sig(value,known_element,m)
            print(m,sig)
            self.add_values(known_element,{"moyenne : "+value:m,"ecart type : "+value:sig})
    def close(self):
        if not(self.aux_loaded=="") :
            self.aux_workbook.close()
            
if __name__=="__main__" :
    import os
    import time
    def get_last_file_of(path):
        maxt=0
        L = os.listdir(path)
        for x in L :
            if x.endswith(".xlsx") :
                maxt = max(os.path.getmtime(os.path.join(path,x)),maxt)
        if not(maxt==0) :
            for x in L :
                if x.endswith(".xlsx") and os.path.getmtime(os.path.join(path,x))==maxt :
                    return os.path.join(path,x)
        return None
    
    path0 = get_last_file_of(os.path.join(os.getcwd(),"OSFI_lien_entre_compteur_et_batiment"))
    path1 = get_last_file_of(os.path.join(os.getcwd(),"OSFI_donnees_du_RT"))
    path2023 = get_last_file_of(os.path.join(os.getcwd(),"OSFI_consommation_mensuelle_2023"))
    path2022 = get_last_file_of(os.path.join(os.getcwd(),"OSFI_consommation_mensuelle_2022"))
    path3 = os.path.join(os.getcwd(),"result.xlsx")
    oad_path = get_last_file_of(os.path.join(os.getcwd(),"OAD_liste_des_batiments_et_de_leurs_surface"))
    t=time.time()
    XL = Excel(path1)
    XL.remove(known_element={"Typologie du bâtiment":["OUVRAGES D'ART RÉSEAUX VOIRIES","MONUMENT ET MÉMORIAL","ESPACE AMÉNAGÉ","ESPACE NATUREL","RÉSEAUX ET VOIRIES"]})
    XL.extract_ending_num({"Code Site":"Code Site RT"})
    XL.import_columns_from(path=path0,links_main_to_imported={"Identifiant du bâtiment":"Identifiant du bâtiment"},cols_to_import={"Fluide":"Fluides relevés pour le bâtiment"},collapsed="concat")
    
    #XL.save(path3[:-5]+"_step1.xlsx")
    # XL = Excel(path3[:-5]+"_step1.xlsx")
    
    XL.import_columns_from(path=path2022,links_main_to_imported={"Identifiant du bâtiment":"Identifiant du bâtiment"},cols_to_import={"Gaz - Consommation":"2022 Gaz - Consommation","Électricité - Consommation":"2022 Électricité - Consommation","Réseau de chaleur - Consommation":"2022 Réseau de chaleur - Consommation","Réseau de froid - Consommation":"2022 Réseau de froid - Consommation","Fioul - Consommation":"2022 Fioul - Consommation","Consommation de granulés de bois":"2022 Consommation de granulés de bois"},where={},collapsed="summ",count_imported=True)
    XL.import_columns_from(path=path2022,links_main_to_imported={"Identifiant du bâtiment":"Identifiant du bâtiment"},cols_to_import={"Gaz - Consommation":"Mai 2022 Gaz - Consommation","Électricité - Consommation":"Mai 2022 Électricité - Consommation","Réseau de chaleur - Consommation":"Mai 2022 Réseau de chaleur - Consommation","Réseau de froid - Consommation":"Mai 2022 Réseau de froid - Consommation","Fioul - Consommation":"Mai 2022 Fioul - Consommation","Consommation de granulés de bois":"Mai 2022 Consommation de granulés de bois"},where={"Date":"2022-05-01"},collapsed="summ",count_imported=False)
    XL.import_columns_from(path=path2022,links_main_to_imported={"Identifiant du bâtiment":"Identifiant du bâtiment"},cols_to_import={"Gaz - Consommation":"Hiver 2022 Gaz - Consommation","Électricité - Consommation":"Hiver 2022 Électricité - Consommation","Réseau de chaleur - Consommation":"Hiver 2022 Réseau de chaleur - Consommation","Réseau de froid - Consommation":"Hiver 2022 Réseau de froid - Consommation","Fioul - Consommation":"Hiver 2022 Fioul - Consommation","Consommation de granulés de bois":"Hiver 2022 Consommation de granulés de bois"},where={"Date":["2022-01-01","2022-02-01","2022-01-01","2022-03-01","2022-04-01","2022-05-01","2022-10-01","2022-11-01","2022-12-01"]},collapsed="summ",count_imported=False)
    XL.import_columns_from(path=path2022,links_main_to_imported={"Identifiant du bâtiment":"Identifiant du bâtiment"},cols_to_import={"Gaz - Consommation":"Ete 2022 Gaz - Consommation","Électricité - Consommation":"Ete 2022 Électricité - Consommation","Réseau de chaleur - Consommation":"Ete 2022 Réseau de chaleur - Consommation","Réseau de froid - Consommation":"Ete 2022 Réseau de froid - Consommation","Fioul - Consommation":"Ete 2022 Fioul - Consommation","Consommation de granulés de bois":"Ete 2022 Consommation de granulés de bois"},where={"Date":["2022-06-01","2022-07-01","2022-08-01","2022-09-01"]},collapsed="summ",count_imported=False)
    
    #XL.save(path3[:-5]+"_step2.xlsx")
    # XL = Excel(path3[:-5]+"_step2.xlsx")
    
    XL.import_columns_from(path=path2023,links_main_to_imported={"Identifiant du bâtiment":"Identifiant du bâtiment"},cols_to_import={"Gaz - Consommation":"2023 Gaz - Consommation","Électricité - Consommation":"2023 Électricité - Consommation","Réseau de chaleur - Consommation":"2023 Réseau de chaleur - Consommation","Réseau de froid - Consommation":"2023 Réseau de froid - Consommation","Fioul - Consommation":"2023 Fioul - Consommation","Consommation de granulés de bois":"2023 Consommation de granulés de bois"},where={},collapsed="summ",count_imported=True)
    XL.import_columns_from(path=path2023,links_main_to_imported={"Identifiant du bâtiment":"Identifiant du bâtiment"},cols_to_import={"Gaz - Consommation":"Mai 2023 Gaz - Consommation","Électricité - Consommation":"Mai 2023 Électricité - Consommation","Réseau de chaleur - Consommation":"Mai 2023 Réseau de chaleur - Consommation","Réseau de froid - Consommation":"Mai 2023 Réseau de froid - Consommation","Fioul - Consommation":"Mai 2023 Fioul - Consommation","Consommation de granulés de bois":"Mai 2023 Consommation de granulés de bois"},where={"Date":"2023-05-01"},collapsed="summ",count_imported=False)
    XL.import_columns_from(path=path2023,links_main_to_imported={"Identifiant du bâtiment":"Identifiant du bâtiment"},cols_to_import={"Gaz - Consommation":"Hiver 2023 Gaz - Consommation","Électricité - Consommation":"Hiver 2023 Électricité - Consommation","Réseau de chaleur - Consommation":"Hiver 2023 Réseau de chaleur - Consommation","Réseau de froid - Consommation":"Hiver 2023 Réseau de froid - Consommation","Fioul - Consommation":"Hiver 2023 Fioul - Consommation","Consommation de granulés de bois":"Hiver 2023 Consommation de granulés de bois"},where={"Date":["2023-01-01","2023-02-01","2023-03-01","2023-04-01","2023-05-01","2023-10-01","2023-11-01","2023-12-01"]},collapsed="summ",count_imported=False)
    XL.import_columns_from(path=path2023,links_main_to_imported={"Identifiant du bâtiment":"Identifiant du bâtiment"},cols_to_import={"Gaz - Consommation":"Ete 2023 Gaz - Consommation","Électricité - Consommation":"Ete 2023 Électricité - Consommation","Réseau de chaleur - Consommation":"Ete 2023 Réseau de chaleur - Consommation","Réseau de froid - Consommation":"Ete 2023 Réseau de froid - Consommation","Fioul - Consommation":"Ete 2023 Fioul - Consommation","Consommation de granulés de bois":"Ete 2023 Consommation de granulés de bois"},where={"Date":["2023-06-01","2023-07-01","2023-08-01","2023-09-01"]},collapsed="summ",count_imported=False)
    
    #XL.save(path3[:-5]+"_step3.xlsx")
    # XL = Excel(path3[:-5]+"_step3.xlsx")
    
    XL.import_columns_from(path=oad_path,links_main_to_imported={"Code bâtiment RT":"Code bât/ter","Code Site RT":"Code Site"},cols_to_import={"Surface de plancher":"SDP RT","SUB":"SUB RT"},collapsed="last")
    
    # XL.save(path3[:-5]+"_step4.xlsx")
    # XL = Excel(path3[:-5]+"_step4.xlsx")
    
    L=XL.get_list_from_col("Code Site",known_element={"Nombre d éléments importés":0})
    for x in L :
        XL.virtual_group_by_sum(known_element={"Code Site":x}, to_sum_name=["Surface au sol","2022 Gaz - Consommation","2022 Électricité - Consommation","2022 Réseau de chaleur - Consommation","2022 Réseau de froid - Consommation","2022 Fioul - Consommation","2022 Consommation de granulés de bois","2023 Gaz - Consommation","2023 Électricité - Consommation","2023 Réseau de chaleur - Consommation","2023 Réseau de froid - Consommation","2023 Fioul - Consommation","2023 Consommation de granulés de bois"],result_prefix="Regroupé par site ",add_count=True)
    #XL = Excel(path3[:-5]+"_step2.xlsx")
    XL.import_columns_from(path=oad_path,links_main_to_imported={"Code bâtiment RT":"Code bât/ter","Code Site RT":"Code Site"},cols_to_import={"Surface de plancher":"SDP RT","SUB":"SUB RT"},collapsed="last")
    L = XL.get_list_from_cols(["Code Site","Code bâtiment RT"],{})
    for [x,y] in L :
        XL.virtual_group_by_sum(known_element={"Code Site":x,"Code bâtiment RT":y}, to_sum_name=["Surface au sol"],result_prefix="Regroupé par bâtiment")
    XL.is_in(["Typologie du bâtiment","Typologie détaillée"],[["BÂT. ENSEIGNEMENT OU SPORT"],
		["BÂTIMENT CULTUREL"],
		["BATIMENT SANITAIRE OU SOCIAL"],
		["BUREAU"],
		["BATIMENT TECHNIQUE","BÂTIMENT TECHNIQUE"],
		["BATIMENT TECHNIQUE","BÂTIMENT SCIENTIFIQUE"],
		["BATIMENT TECHNIQUE","CENTRE DE RECHERCHE OU D'ESSAI"],
		["BATIMENT TECHNIQUE","DÉPÔT D'ARCHIVES"],
		["BATIMENT TECHNIQUE","CENTRE INFORMATIQUE"],
		["BATIMENT TECHNIQUE","POSTE DE COMMANDEMENT"]],"dans le périmètre choisi")
    XL.get_stat_by_element(col = "Typologie du bâtiment", value = "Surface au sol")
    XL.save(path3)
    XL.close()
    print(str(int(time.time()-t))+" secondes")
    
    
        
        
            
    
        