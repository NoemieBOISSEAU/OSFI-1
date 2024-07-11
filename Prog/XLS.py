# -*- coding: utf-8 -*-
"""
Created on Tue Jan 23 19:38:37 2024

@author: sacha
"""
import os
import openpyxl
def float_(s) :
    if s==None :
        return False,0
    if isinstance(s,float) :
        return True,s
    if isinstance(s,int) :
        return True,s
    while (((s.startswith('"'))and(s.endswith('"')))or((s.startswith("'"))and(s.endswith("'")))) :
        s=s[1:-1]
    if s == "" :
        return False,0
    try :
        return True,float(s)
    except :
        try :
            return True,float(s.replace(",","."))
        except :
            return True,s
def int_(s):
    if s==None :
        return False,0
    if isinstance(s,float) :
        return True,int(s)
    if isinstance(s,int) :
        return True,s
    while (((s.startswith('"'))and(s.endswith('"')))or((s.startswith("'"))and(s.endswith("'")))) :
        s=s[1:-1]
    if s == "" :
        return False,0
    try :
        return True,int(s)
    except :
        try :
            return True,int(float(s))
        except :
            try :
                return True,int(float(s.replace(",",".")))
            except :
                return True,s
def str_(s):
    if s==None :
        return False,""
    if isinstance(s,float) :
        return True,str(s)
    if isinstance(s,int) :
        return True,str(s)
    while (((s.startswith('"'))and(s.endswith('"')))or((s.startswith("'"))and(s.endswith("'")))) :
        s=s[1:-1]
    return True,s
class Excel :
    def __init__(self,path=""):
        print("Start loading")
        self.L_seps=["",'"','""','"""',"'","''","'''"]
        self.loaded=True
        try :
            self.workbook=openpyxl.load_workbook(filename=path)
        except :
            self.loaded=False
        print("End loading")
    def index(self,name):
        Lnames=[name]
        # flag2,flag3=True,True
        # try : 
        #     name2 = name.encode("cp1252").decode("utf-8")
        # except :
        #     flag2=False
        # try : 
        #     name3 = name.encode("utf-8").decode("cp1252")
        # except :
        #     flag3=False
        # if flag2 :
        #     Lnames.append(name2)
        # if flag3 :
        #     Lnames.append(name3)
        flag=False
        if self.loaded :
            for sheet in self.workbook.sheetnames :
                row,col,flag_col=1,1,True
                while flag_col and not(flag) :
                    txt=self.workbook[sheet].cell(row,col).value
                    if txt==None :
                        flag_col=False
                    else :
                        for sep in self.L_seps :
                            if txt==sep+""+sep :
                                flag_col=False
                                break
                            elif txt==sep+Lnames[0]+sep :
                                flag=True
                                break
                            elif len(Lnames)>1 and txt==sep+Lnames[1]+sep :
                                flag=True
                                break
                            elif len(Lnames)>2 and txt==sep+Lnames[2]+sep :
                                flag=True
                                break
                    col+=1
                if flag :
                    break
        if flag :
            print(name,self.workbook[sheet].cell(row,col-1).value)
            return sheet,col-1
        print(False)
        return "",0
    def get_lists(self,L_names):
        if self.loaded :
            Titles=[]
            for x in L_names :
                print(x)
                if isinstance(x,list) :
                    print("list)")
                    found_good=False
                    for _x_ in x :
                        _type_=str_
                        if isinstance(_x_,dict) :
                            sheet,col=self.index(_x_["name"])
                            if "type" in _x_ :
                                if _x_["type"]=="str" :
                                    _type_=str_
                                elif _x_["type"]=="float" :
                                    _type_=float_
                                elif _x_["type"]=="int" :
                                    _type_=int_
                        elif isinstance(_x_,str) :
                            sheet,col=self.index(_x_)
                        if not(sheet=="") :
                            Titles.append([sheet,col,_type_])
                            found_good=True
                            break
                    if not(found_good) :
                        Titles.append("",0,_type_)
                else :
                    print("direct")
                    _type_=str_
                    if isinstance(x,dict) :
                        sheet,col=self.index(x["name"])
                        if "type" in x :
                            if x["type"]=="str" :
                                _type_=str_
                            elif x["type"]=="float" :
                                _type_=float_
                            elif x["type"]=="int" :
                                _type_=int_
                    elif isinstance(x,str) :
                        sheet,col=self.index(x)
                    Titles.append([sheet,col,_type_])
            L,row,flag_void=[],2,True
            while flag_void :
                flag_void=False
                line=[]
                for [sheet,col,_type_] in Titles :
                    if sheet=="" :
                        line.append(_type_(""))
                    else :
                        flag,val=_type_(self.workbook[sheet].cell(row,col).value)
                        line.append(val)
                        flag_void=(flag_void or flag)
                row+=1
                if flag_void :
                    L.append(line)
            return L
        return []
    def get_typed_lists(self,L_names,_type_="float"):
        L=[]
        for name in L_names :
            if isinstance(name,list) :
                _l_=[]
                for x in name :
                    _l_.append({"name":x,"type":_type_})
                L.append(_l_)
            else :
                L.append({"name":name,"type":_type_})
        return self.get_lists(L_names=L)
    def close(self):
        if self.loaded :
            self.workbook.close()
            self.loaded=False
if __name__=="__main__" :
    situation=["Mauvaise identification du bâtiment","Surface négative","DJU négatif","Au moins une des consommation est négative","La consommation totale est nulle","La consommation totale est improbable","Tout semble correct"]
    Lx=["Surface totale du bâtiment","Degrés jours unifiés"]
    Lmeta=["Code bâtiment RT","Nom du bâtiment","Code Site","Nom du site","Année","Etat du bâtiment"]
    Ly=["Consommation d'électricité (kWh)","Consommation de gaz (kWh PCS)",
        "Consommation du réseau de chaud (kWh)","Consommation du réseau de froid (kWh)",
        "Consommation de fioul (kWh PCS)","Consommation de granulés de bois (kWh)"]
    filename="Consommations annuelles des équipements.xlsx"
    path=os.getcwd()
    path=os.path.join(path,filename)
    XL = Excel(path) 
    X=XL.get_typed_lists([Lmeta[0]]+Lx,"float")
    for i in range(len(X)) :
        del X[i][0]
    META=XL.get_lists(Lmeta)
    Y=XL.get_typed_lists([Lmeta[0]]+Ly,"float")
    for i in range(len(Y)) :
        del Y[i][0]
    XL.close()
    idp=[min(len(_meta_[0]),len(_meta_[2]))>2 for _meta_ in META]
    ye,yt,yp=[_y_[0] for _y_ in Y],[sum(_y_) for _y_ in Y],[(min(_y_)>=0) for _y_ in Y]
    sp,djup=[(_x_[0]>=0) for _x_ in X],[(_x_[1]>=0) for _x_ in X]
    print([x[0] for x in X])
    # for i in range(len(META)) :
    #     print(str(i)+" : "+str(X[i])+str(META[i])+str(Y[i])+str(yt[i]))
        
        