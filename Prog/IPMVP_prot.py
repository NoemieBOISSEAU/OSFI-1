# -*- coding: utf-8 -*-
"""
Created on Tue Jan  7 17:30:42 2025

@author: sacha.mailler
"""
from openpyxl import load_workbook
import os
import statsmodels.api as sm

class IPMVP :
    def __init__(self,config):
        self.talon,self.chau,self.froi,self.use = 0,0,0,0
        self.ref_datas = []
        self.__get_regression_meta_data(config)
        print(self.y)
        print(self.X)
        print(self.Ids)
    def add_ref(self,data) :
        def get_element_of(L):
            count = 0
            for x in L :
                if x in data :
                    count+=1
                if count>1 :
                    raise Exception("Une même variable peut correspondre à plusisurs éléments de vos données")
                    break
            if count == 0 :
                raise Exception("Aucun élément trouvé")
            elif count==1 :
                for x in L :
                    if x in data :
                        return data[x]
        y,X=get_element_of(self.y),[]
        for x in self.X :
            X.append(get_element_of(self.X[x]))
        self.ref_datas.push({"y":y,"X":X})
    def __get_regression_meta_data(self,path):
        self.y,self.X,self.Ids,flag,i={},{},[],True,0
        workbook = load_workbook(path)
        sheet = workbook["IPMVP"]
        def extract_col(j):
            k,List_of_values = 3,[]
            print(j,k)
            print(sheet.cell(row=k,column=j).value)
            while not(sheet.cell(row=k,column=j).value in ["",None]) :
                List_of_values.append(sheet.cell(row=k,column=j).value)
                k+=1
            return List_of_values
        while flag :
            i+=1
            a1,a2 = sheet.cell(row=1,column=i).value,sheet.cell(row=2,column=i).value
            if a1 in ["",None] :
                flag=False
            elif a1=="y" :
                if len(self.y)>+1 :
                    raise Exception("Deux colonnes distinctes définissent y")
                else :
                    self.y=extract_col(i)
            elif a1.startswith("x"):
                if a2 in self.X :
                    raise Exception("Un même nom est utilisé pour plus d'une variable distincte")
                else :
                    self.X[a2]=extract_col(i)
            elif a1=="id" :
                if len(self.Ids)>0 :
                    raise Exception("Une autre colonne identifiant à déjà été lue")
                else :
                    self.Ids=extract_col(i)
            else :
                raise Exception("Type d'élément non prévu")

if __name__=="__main__" :
    print(__file__)
    config = os.path.join(os.path.join(os.path.dirname(os.path.dirname(__file__)),"Datas"),"meta_data.xlsx")
    math = IPMVP(config)
    