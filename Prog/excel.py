# -*- coding: utf-8 -*-
"""
Created on Wed Jan 22 14:28:40 2025

@author: sacha.mailler
"""

import openpyxl
import random

class Excel :
    def __init__(self):
        self.path=None
        self.workbok = None
        self.L_names = []
        self.L_values = []
    def read(self,path=None):
        if not(self.path==None) :
            self.workbook.close()
        if not(path==None) :
            self.path = path
        self.workbook = openpyxl.load_workbook(self.path)
    def close(self):
        if not(self.path==None) :
            self.workbook.close()
    def __get_max_col(self,max_void=5):
        col,count_void=0,0
        while count_void<max_void :
            col+=1
            if self.workbook.active.cell(row=1,column=col).value in [None,""] :
                count_void+=1
            else :
                count_void = 0
        return col-count_void
    def __get_max_line(self,max_col=None,max_void = 5):
        if max_col==None :
            max_col = self.__get_max_col(max_void)
        def is_void_line(line):
            flag = True
            for i in range(max_col) :
                flag = self.workbook.active.cell(row=line,column=i+1).value in [None,""]
                if not(flag) :
                    break
            return flag
        line,count_void=1,0
        while count_void<max_void :
            line+=1
            if is_void_line(line) :
                count_void+=1
            else :
                count_void = 0
        return line-count_void
                
    def __index(self,name, simplified=False,max_col=None):
        if max_col==None :
            max_col  = self.__get_max_col()
        to_replace = "äâàABCDêéèëEFGHIîïJKLMNOôöPQRSTUûüùVWXYZ"
        replacer   = "aaaabcdeeeeefghiiijklmnooopqrstuuuuvwxyz"
        def simplify(word):
            if simplified :
                result=""
                for i in range(len(word)) :
                    if word[i] in replacer :
                        result+=word[i]
                    elif word[i] in to_replace :
                        result+=replacer[to_replace.index(word[i])]
                return result
            else :
                return word
        for i in range(max_col) :
            if simplify(self.workbook.active.cell(row=1,column=i+1).value)==simplify(name) :
                return i
        return -1
    def load(self,path,dict_of_names,rad="",term=""):
        flag = True
        for x in dict_of_names :
            if rad+x+term in self.L_names :
                flag = False
                raise Exception("Error : La grandeur"+rad+x+term+"à charger à déjà été chargée ! ")
        if flag :
            self.read(path)
            col_num = self.__get_max_col()
            line_num = self.__get_max_line()
            for name in dict_of_names :
                if isinstance(dict_of_names[name],list) :
                    for s_name in dict_of_names[name] :
                        ind = self.__index(s_name,simplified=False)
                        if not(ind==-1) :
                            break
                    if ind == -1 :
                        for s_name in dict_of_names[name] :
                            ind = self.__index(s_name,simplified=True)
                            if not(ind==-1) :
                                break
                else :
                    ind = self.__index(s_name,simplified=False)
                    if ind == -1 :
                        ind = self.__index(s_name,simplified=True)
                if ind==-1 :
                    raise Exception("Error : Aucune des grandeurs "+str(dict_of_names[name])+" n'a pas été trouvée dans le fichier"+self.path)
                else :
                    self.L_names.append(rad+name+term)
                    for j in range(line_num-1):
                        if len(self.L_values)<=j :
                            self.L_values.append([])
                            for i in range(len(self.L_names)-1) :
                                self.L_values[-1].append(None)
                        self.L_values[j].append(self.workbook.active.cell(row=j+2,column=ind+1).value)
            self.close()
    def display(self):
        L_size,L=[],[]
        for i in range(100) :
            L.append(random.randint(0,len(self.L_values)))
        L.sort()
        for j in range(len(self.L_names)) :
            L_size.append(len(self.L_names[j]))
        for i in range(len(L)) :
            for j in range(len(self.L_names)) :
                L_size[j]=max(L_size[j],len(str(self.L_values[L[i]][j])))
        txt=""
        for j in range(len(self.L_names)) :
            txt+=self.L_names[j]+" "*(L_size[j]-len(self.L_names[j]))+"|"
        print(txt)
        for i in range(len(L)) :
            txt=""
            for j in range(len(self.L_names)) :
                txt+=str(self.L_values[L[i]][j])+" "*(L_size[j]-len(str(self.L_values[L[i]][j])))+"|"
            print(txt)
    def __exit__(self,exc_type,exc_value,traceback):
        if not(self.path==None) :
            self.workbook.close()
        print("Type : "+str(exc_type))
        print("Value : "+str(exc_value))
        print("Traceback : "+str(traceback))
        
if __name__=="__main__" :
    import os
    path = os.path.join(os.getcwd(),"OSFI_consommation_mensuelle_2023")
    IDS = {"ID":["Identifiant du bâtiment"],"DATE":["Date"]}
    Consos = {"DJU":["Degrés-jours (DJ) de chauffage"],
             "DJF":["Degrés-jours (DJ) de refroidissement"],
             "fioul":["Fioul - Consommation"],
             "gaz":["Gaz - Consommation"],
             "rcu":["Réseau de chaleur - Consommation"],
             "bois":["Consommation de granulés de bois"],
             "froid":["Réseau de froid - Consommation"],
             "elec":["Électricité - Consommation"]}
    Datas = {"code bat RT":["Code bâtiment RT"],
             "code site RT":["Code Site"],
             "surface":["Surface au sol"],
             "typologie 1":["Typologie du bâtiment"],
             "typologie 2": ["Typologie détaillée"],}
    E = Excel()
    E.load(os.path.join(path,os.listdir(path)[0]),IDS)
    E.display()
    E.load(os.path.join(path,os.listdir(path)[0]),Datas)
    E.display()