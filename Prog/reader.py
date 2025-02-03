# -*- coding: utf-8 -*-
"""
Created on Tue Jan 28 15:39:48 2025

@author: sacha.mailler
"""
class VALUE : 
    def __init__(self,val):
        if str(val) in ["NaN","nan",""] :
            self.value=None
        else :
            self.value = val
class sub_pandas_object :
    def __init__(self,DF) :
        self.df = DF
    def cell(self,row=1,column=1):
        if row == 1 :
            try :
                #print(str(column)+" : "+str(self.df.columns[column-1]))
                return VALUE(self.df.columns[column-1])
            except :
                return VALUE(None)
            
        else :
            try:
                #print(str(row-1)+", "+str(self.df.columns[column-1])+" : "+self.df.loc[row-1][self.df.columns[column-1]])
                return VALUE(self.df.loc[row-1][self.df.columns[column-1]])
            except :
                return VALUE(None)
    def close(self):
        del self.df
        del self
class pandas_object :
    def __init__(self,df) :
        self.active= sub_pandas_object(df)
    def close(self):
        self.active.close()
        del self
class basic_reader :
    def __init__(self):
        try :
            import openpyxl
            self.openpyxl  =openpyxl
        except :
            self.openpyxl = None
        self.pandas=None
        # try :
        #     import pandas
        #     self.pandas = pandas
        # except :
        #     self.pandas = None
    def load_workbook(self,path,read_only=False):
        if self.pandas == None :
            if self.openpyxl == None :
                raise Exception("Error : Aucun des modules de lecture n'a été chargé")
            else :
                return self.openpyxl.load_workbook(path)
        else :
            return pandas_object(self.pandas.read_excel(path))
    def Workbook(self):
        if self.openpyxl==None :
            raise Exception("Error : Pour l'instant le module openpyxl est nécessaire pour enregistrer les fichiers")
        else :
            return self.openpyxl.Workbook()